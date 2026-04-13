# -*- coding: utf-8 -*-
"""
PROGRAMME : rechercheV (Equivalent à une rechercheV pour .csv)
VERSION : 6.0
MODIFICATIONS :
- Alignement UI strict (Clé Liaison / Nom Fichier / Sources).
- Zone de commentaire multiligne persistante dans la config.
- Bilan détaillé avec pourcentages dans les logs.
- Internationalisation : toutes les chaînes externalisées (strings_fr.py / strings_en.py).
- Sélecteur de langue FR / EN dans l'interface.
"""

import tkinter as tk
from tkinter import ttk, filedialog
import json
import os
import csv
from datetime import datetime
import i18n

CONFIG_FILE = "config_rechercheV.json"


def _load_lang_from_config():
    """Lit la préférence de langue dans la config avant de construire l'UI."""
    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            c = json.load(f)
        i18n.set_lang(c.get("lang", "fr"))
    except Exception:
        pass  # langue par défaut : FR


_load_lang_from_config()
S = i18n.S  # référence au dict de la langue active (figée au démarrage)


class AppRechercheV:
    def __init__(self, root):
        self.root = root
        self.root.title(S["WINDOW_TITLE"])
        self.root.geometry("1100x900")
        self.root.minsize(900, 700)

        self.mapping_rows = []
        self.create_rejet_file = tk.BooleanVar(value=False)
        self.label_width = 30

        # --- Barre de langue (tout en haut) ---
        self._setup_lang_bar()

        self.pane_config = ttk.LabelFrame(self.root, text=S["BLOC1_TITLE"], padding=15)
        self.pane_config.pack(side="top", fill="x", padx=20, pady=5)
        self.setup_bloc1()

        self.mapping_group = ttk.LabelFrame(self.root, text=S["BLOC2_TITLE"], padding=15)
        self.mapping_group.pack(side="top", fill="x", padx=20, pady=5)
        self.setup_bloc2_canvas()

        self.pane_action = ttk.LabelFrame(self.root, text=S["BLOC3_TITLE"], padding=15)
        self.pane_action.pack(side="top", fill="both", expand=True, padx=20, pady=(5, 10))
        self.setup_bloc3_action()

        self.load_config()

    # ------------------------------------------------------------------
    # Sélecteur de langue
    # ------------------------------------------------------------------

    def _setup_lang_bar(self):
        """Deux petits boutons FR / EN en haut à droite."""
        lang_bar = ttk.Frame(self.root)
        lang_bar.pack(side="top", fill="x", padx=12, pady=(3, 0))

        active = i18n.get_lang()
        font_btn = ("Segoe UI", 8, "bold")

        btn_en = tk.Button(
            lang_bar, text="EN", font=font_btn, width=3,
            relief="sunken" if active == "en" else "raised",
            cursor="hand2",
            command=lambda: self._switch_lang("en"),
        )
        btn_en.pack(side="right", padx=(2, 10))

        btn_fr = tk.Button(
            lang_bar, text="FR", font=font_btn, width=3,
            relief="sunken" if active == "fr" else "raised",
            cursor="hand2",
            command=lambda: self._switch_lang("fr"),
        )
        btn_fr.pack(side="right", padx=2)

    def _switch_lang(self, lang):
        """Sauvegarde la préférence de langue et redémarre l'application."""
        if lang == i18n.get_lang():
            return
        # Lire la config existante pour ne pas l'écraser
        cfg = {}
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                cfg = json.load(f)
        except Exception:
            pass
        cfg["lang"] = lang
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=4, ensure_ascii=False)
        # Relancer l'application avec la nouvelle langue
        import subprocess, sys
        subprocess.Popen([sys.executable] + sys.argv)
        self.root.after(100, self.root.destroy)

    # ------------------------------------------------------------------
    # Bloc 1 — Sélection des fichiers et paramètres
    # ------------------------------------------------------------------

    def setup_bloc1(self):
        self.path_f1 = self.create_file_selection(self.pane_config, S["LABEL_FILE1"])
        self.path_f2 = self.create_file_selection(self.pane_config, S["LABEL_FILE2"])

        # Fichier résultat : chemin complet (dossier + nom) sur une seule ligne
        row_res = ttk.Frame(self.pane_config)
        row_res.pack(fill="x", pady=5)
        ttk.Label(row_res, text=S["LABEL_RESULT"], width=self.label_width).pack(side="left")
        self.path_result = tk.StringVar(value="")
        ttk.Entry(row_res, textvariable=self.path_result).pack(side="left", fill="x", expand=True, padx=5)
        ttk.Button(row_res, text="...", width=3, command=self._choisir_fichier_resultat).pack(side="left", padx=2)

        # Format + rejet sur la même ligne
        row_fmt = ttk.Frame(self.pane_config)
        row_fmt.pack(fill="x", pady=3)
        ttk.Label(row_fmt, text=S["LABEL_FORMAT"], width=self.label_width).pack(side="left")
        self.out_format = tk.StringVar(value="csv")
        ttk.Radiobutton(row_fmt, text=S["FMT_CSV"], variable=self.out_format,
                        value="csv", command=self._on_format_change).pack(side="left", padx=5)
        ttk.Radiobutton(row_fmt, text=S["FMT_XLSX"], variable=self.out_format,
                        value="xlsx", command=self._on_format_change).pack(side="left", padx=5)
        ttk.Checkbutton(row_fmt, text=S["CHK_REJET"],
                        variable=self.create_rejet_file).pack(side="left", padx=20)

        row_keys = ttk.Frame(self.pane_config)
        row_keys.pack(fill="x", pady=5)
        ttk.Label(row_keys, text=S["LABEL_KEY1"], width=self.label_width).pack(side="left")
        self.key_f1 = tk.StringVar(value="id")
        ttk.Entry(row_keys, textvariable=self.key_f1, width=30).pack(side="left", padx=5)

        ttk.Label(row_keys, text=S["LABEL_KEY2"], padding=(20, 0, 5, 0)).pack(side="left")
        self.key_f2 = tk.StringVar(value="id")
        ttk.Entry(row_keys, textvariable=self.key_f2, width=30).pack(side="left", padx=5)

        self.case_sensitive = tk.BooleanVar(value=False)
        ttk.Checkbutton(row_keys, text=S["CHK_CASE"], variable=self.case_sensitive).pack(side="left", padx=15)

        ttk.Label(self.pane_config, text=S["LABEL_COMMENT"]).pack(anchor="w", pady=(10, 0))
        self.comment_area = tk.Text(self.pane_config, height=4, font=("Segoe UI", 9))
        self.comment_area.pack(fill="x", pady=5)

    # ------------------------------------------------------------------
    # Bloc 2 — Mapping des colonnes
    # ------------------------------------------------------------------

    def setup_bloc2_canvas(self):
        container = ttk.Frame(self.mapping_group)
        container.pack(fill="both", expand=True)
        self.main_canvas = tk.Canvas(container, highlightthickness=0, height=180)
        self.scrollbar = ttk.Scrollbar(container, orient="vertical", command=self.main_canvas.yview)
        self.mapping_area = ttk.Frame(self.main_canvas)
        self.main_canvas.bind('<Configure>', lambda e: self.main_canvas.itemconfig(self.canvas_window, width=e.width))
        self.mapping_area.bind("<Configure>", lambda e: self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all")))
        self.canvas_window = self.main_canvas.create_window((0, 0), window=self.mapping_area, anchor="nw")
        self.main_canvas.configure(yscrollcommand=self.scrollbar.set)
        self.scrollbar.pack(side="right", fill="y")
        self.main_canvas.pack(side="left", fill="both", expand=True)
        btn_bar = ttk.Frame(self.mapping_group)
        btn_bar.pack(fill="x", pady=(10, 0))
        ttk.Button(btn_bar, text=S["BTN_ADD_ROW"], command=self.add_mapping_row).pack(side="left", padx=5)
        ttk.Button(btn_bar, text=S["BTN_SAVE"], command=self.save_config).pack(side="right", padx=5)

    # ------------------------------------------------------------------
    # Bloc 3 — Traitement et statistiques
    # ------------------------------------------------------------------

    def setup_bloc3_action(self):
        btn_bar = ttk.Frame(self.pane_action)
        btn_bar.pack(fill="x", pady=(0, 8))
        self.run_btn = tk.Button(btn_bar, text=S["BTN_RUN"],
                                 bg="#27ae60", fg="white", font=("Helvetica", 12, "bold"),
                                 command=self.run_process)
        self.run_btn.pack(side="left", fill="x", expand=True)
        tk.Button(btn_bar, text=S["BTN_CLEAR_LOG"],
                  command=lambda: self.log_area.delete("1.0", tk.END),
                  bg="#475569", fg="white", font=("Helvetica", 12), relief="flat", padx=12).pack(side="left", padx=(8, 0))
        self.log_area = tk.Text(self.pane_action, bg="#1e293b", fg="#38bdf8", font=("Consolas", 10))
        self.log_area.pack(fill="both", expand=True)

    # ------------------------------------------------------------------
    # Helpers UI
    # ------------------------------------------------------------------

    def create_file_selection(self, parent, label, is_dir=False):
        frame = ttk.Frame(parent)
        frame.pack(fill="x", pady=5)
        ttk.Label(frame, text=label, width=self.label_width).pack(side="left")
        v = tk.StringVar()
        ttk.Entry(frame, textvariable=v).pack(side="left", fill="x", expand=True, padx=5)
        def br():
            p = filedialog.askdirectory() if is_dir else filedialog.askopenfilename(
                filetypes=[
                    (S["FILETYPES_SUPPORTED"], "*.csv *.xlsx *.xls"),
                    (S["FILETYPES_CSV"],        "*.csv"),
                    (S["FILETYPES_EXCEL"],      "*.xlsx *.xls"),
                ])
            if p:
                v.set(p)
        ttk.Button(frame, text="...", width=3, command=br).pack(side="left", padx=2)
        return v

    def add_mapping_row(self, val_f1="", val_f2=""):
        f = ttk.Frame(self.mapping_area)
        f.pack(fill="x", expand=True, pady=2)
        f.columnconfigure(0, weight=1)
        f.columnconfigure(1, weight=1)
        e_f1 = ttk.Entry(f); e_f1.insert(0, val_f1); e_f1.grid(row=0, column=0, sticky="ew", padx=5)
        e_f2 = ttk.Entry(f); e_f2.insert(0, val_f2); e_f2.grid(row=0, column=1, sticky="ew", padx=5)
        self.mapping_rows.append((e_f1, e_f2))
        tk.Button(f, text="X", fg="red", relief="flat", command=lambda: f.destroy()).grid(row=0, column=2, padx=5)
        self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))

    def get_active_mapping(self, only_complete=False):
        """Retourne les paires (f1, f2) des lignes de mapping encore vivantes.
        only_complete=True : exclut les lignes où la source F2 est vide (pour le traitement).
        only_complete=False : conserve toutes les lignes vivantes (pour la sauvegarde).
        """
        result = []
        for e_f1, e_f2 in self.mapping_rows:
            try:
                v1 = e_f1.get().strip()
                v2 = e_f2.get().strip()
                if only_complete and not v2:
                    continue
                result.append((v1, v2))
            except tk.TclError:
                pass  # widget détruit via le bouton X
        return result

    def log(self, t):
        self.log_area.insert(tk.END, t + "\n")
        self.log_area.see(tk.END)
        self.root.update_idletasks()

    # ------------------------------------------------------------------
    # Gestion du fichier résultat
    # ------------------------------------------------------------------

    def _choisir_fichier_resultat(self):
        """Ouvre un dialogue de sauvegarde pour choisir dossier + nom du fichier résultat."""
        fmt = self.out_format.get()
        if fmt == "xlsx":
            ftypes = [(S["FILETYPES_EXCEL"], "*.xlsx"), (S["FILETYPES_ALL"], "*.*")]
            defext = ".xlsx"
        else:
            ftypes = [(S["FILETYPES_CSV"], "*.csv"), (S["FILETYPES_ALL"], "*.*")]
            defext = ".csv"
        chemin = filedialog.asksaveasfilename(
            filetypes=ftypes,
            defaultextension=defext,
            initialfile=f"{S['INIT_FILE_RESULT']}{defext}",
        )
        if chemin:
            self.path_result.set(chemin)

    def _on_format_change(self):
        """Met à jour l'extension du fichier résultat quand le format change."""
        chemin = self.path_result.get()
        if chemin:
            base = os.path.splitext(chemin)[0]
            ext = ".xlsx" if self.out_format.get() == "xlsx" else ".csv"
            self.path_result.set(base + ext)

    # ------------------------------------------------------------------
    # Lecture de fichiers
    # ------------------------------------------------------------------

    def _est_excel(self, chemin):
        return os.path.splitext(chemin)[1].lower() in ('.xlsx', '.xls')

    def _lire_header(self, chemin):
        """Retourne la liste des noms de colonnes (header) sans charger toutes les données."""
        if self._est_excel(chemin):
            try:
                import openpyxl
            except ImportError:
                raise Exception(S["ERR_OPENPYXL"])
            wb = openpyxl.load_workbook(chemin, read_only=True, data_only=True)
            ws = wb.active
            header = [self.nettoyer_cellule(c.value) for c in next(ws.iter_rows())]
            wb.close()
            return header
        else:
            enc = self._detecter_encodage(chemin)
            with open(chemin, 'r', encoding=enc, errors='replace') as f:
                return next(csv.reader(f, delimiter=';'))

    def _lire_lignes_excel(self, chemin, cols_utiles=None):
        """Générateur : lit un Excel ligne par ligne, retourne des dict {col: valeur}.
        Si cols_utiles fourni, ne retourne que ces colonnes (+ toutes si None)."""
        try:
            import openpyxl
        except ImportError:
            raise Exception(S["ERR_OPENPYXL"])
        wb = openpyxl.load_workbook(chemin, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows()
        headers = [self.nettoyer_cellule(c.value) for c in next(rows)]
        for row in rows:
            vals = [self.nettoyer_cellule(c.value) for c in row]
            d = dict(zip(headers, vals))
            if cols_utiles is not None:
                yield {k: d.get(k, "") for k in cols_utiles}
            else:
                yield d
        wb.close()

    def _detecter_encodage(self, chemin_fichier):
        """Détecte l'encodage d'un fichier CSV sans le modifier."""
        if self._est_excel(chemin_fichier):
            return None  # non applicable
        with open(chemin_fichier, 'rb') as f:
            return 'utf-8-sig' if f.read(3).startswith(b'\xef\xbb\xbf') else 'cp1252'

    def find_col(self, col_name, header_list):
        t = str(col_name).strip().lower()
        for h in header_list:
            if h.lower() == t:
                return h
        return None

    def nettoyer_cellule(self, texte):
        if texte is None:
            return ""
        return str(texte).replace('\r\n', ' ').replace('\r', ' ').replace('\n', ' ').strip()

    # ------------------------------------------------------------------
    # Archivage et nettoyage
    # ------------------------------------------------------------------

    def archiver_si_existe(self, chemin_complet):
        if os.path.exists(chemin_complet):
            mtime = os.path.getmtime(chemin_complet)
            date_str = datetime.fromtimestamp(mtime).strftime('%Y%m%d_%H%M%S')
            nom_base, ext = os.path.splitext(chemin_complet)
            nouveau_nom = f"{nom_base}_{date_str}{ext}"
            os.rename(chemin_complet, nouveau_nom)
            self.log(S["LOG_ARCHIVED"].format(name=os.path.basename(nouveau_nom)))

    def nettoyer_retours_chariot(self, chemin_fichier):
        """Répare les retours chariot embarqués dans les cellules d'un fichier SOURCE CSV.
        Stratégie : recollement des lignes trop courtes (fausse coupure due à un RC).
        IGNORÉ pour les fichiers Excel (.xlsx, .xls) qui sont des formats binaires.
        NE PAS appeler sur les fichiers résultat (déjà écrits proprement par csv.DictWriter)."""
        if not os.path.exists(chemin_fichier):
            return
        if os.path.splitext(chemin_fichier)[1].lower() in ('.xlsx', '.xls'):
            return
        with open(chemin_fichier, 'rb') as f:
            contenu_bin = f.read()

        bom = b'\xef\xbb\xbf'
        a_bom = contenu_bin.startswith(bom)
        corps = contenu_bin[len(bom):] if a_bom else contenu_bin

        lignes = corps.replace(b'\r\n', b'\n').replace(b'\r', b'\n').split(b'\n')
        while lignes and lignes[-1] == b'':
            lignes.pop()
        if not lignes:
            return
        lignes = [l for l in lignes if l.strip()]
        if not lignes:
            return

        nb_cols = lignes[0].count(b';') + 1
        reparees = [lignes[0]]
        nb_rc = 0
        i = 1
        while i < len(lignes):
            ligne = lignes[i]
            while ligne.count(b';') + 1 < nb_cols and i + 1 < len(lignes):
                i += 1
                ligne = ligne + b' ' + lignes[i]
                nb_rc += 1
            reparees.append(ligne)
            i += 1

        if nb_rc > 0:
            propre = b'\n'.join(reparees) + b'\n'
            with open(chemin_fichier, 'wb') as f:
                if a_bom:
                    f.write(bom)
                f.write(propre)
            self.log(S["LOG_RC_REPAIRED"].format(count=nb_rc, name=os.path.basename(chemin_fichier)))

    def convertir_en_ansi(self, chemin_fichier):
        if os.path.exists(chemin_fichier):
            with open(chemin_fichier, 'r', encoding='utf-8', errors='ignore') as f_in:
                contenu = f_in.read()
            with open(chemin_fichier, 'w', encoding='cp1252', errors='replace') as f_out:
                f_out.write(contenu)

    # ------------------------------------------------------------------
    # Écriture Excel
    # ------------------------------------------------------------------

    def _ecrire_xlsx(self, chemin, headers, lignes_ok, lignes_rej=None):
        """Écrit le résultat en Excel :
        - feuille résultat avec header gras + fond gris
        - feuille rejet si lignes_rej fourni
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise Exception(S["ERR_OPENPYXL"])

        wb = openpyxl.Workbook()

        def ecrire_feuille(ws, headers, lignes):
            fill = PatternFill("solid", fgColor="D9D9D9")
            font = Font(bold=True)
            ws.append(headers)
            for cell in ws[1]:
                cell.font = font
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center")
            for row in lignes:
                ws.append([row.get(h, "") for h in headers])
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        ws_res = wb.active
        ws_res.title = S["SHEET_RESULT"]
        ecrire_feuille(ws_res, headers, lignes_ok)

        if lignes_rej is not None:
            ws_rej = wb.create_sheet(S["SHEET_REJET"])
            ecrire_feuille(ws_rej, headers, lignes_rej)

        wb.save(chemin)

    # ------------------------------------------------------------------
    # Traitement principal
    # ------------------------------------------------------------------

    def run_process(self):
        self.log(S["LOG_INIT"])
        tmp_f1 = tmp_f2 = None
        try:
            f1_path  = os.path.abspath(self.path_f1.get())
            f2_path  = os.path.abspath(self.path_f2.get())
            res_path = os.path.abspath(self.path_result.get())
            out_name = os.path.basename(res_path)
            out_dir  = os.path.dirname(res_path)

            if not f1_path or not os.path.isfile(f1_path):
                raise Exception(S["ERR_F1_NOT_FOUND"])
            if not f2_path or not os.path.isfile(f2_path):
                raise Exception(S["ERR_F2_NOT_FOUND"])
            if not res_path.strip():
                raise Exception(S["ERR_RESULT_EMPTY"])
            if not out_dir or not os.path.isdir(out_dir):
                raise Exception(S["ERR_OUT_DIR"].format(dir=out_dir))

            active_mapping = self.get_active_mapping(only_complete=True)
            if not active_mapping:
                raise Exception(S["ERR_NO_MAPPING"])

            # Copies de travail pour nettoyage RC (CSV uniquement, originaux jamais modifiés)
            import shutil
            def _copie_travail(src):
                if self._est_excel(src):
                    return src  # Excel lu directement, pas de nettoyage RC
                tmp = src + "._tmp_rv"
                shutil.copy2(src, tmp)
                self.nettoyer_retours_chariot(tmp)
                return tmp

            tmp_f1 = _copie_travail(f1_path)
            tmp_f2 = _copie_travail(f2_path)

            # Lecture des headers
            h1 = self._lire_header(tmp_f1)
            h2 = self._lire_header(tmp_f2)
            k1 = self.find_col(self.key_f1.get(), h1)
            k2 = self.find_col(self.key_f2.get(), h2)

            if not k1:
                raise Exception(S["ERR_KEY1_MISSING"].format(key=self.key_f1.get()))
            if not k2:
                raise Exception(S["ERR_KEY2_MISSING"].format(key=self.key_f2.get()))

            final_mapping = []
            for target_f1, source_f2 in active_mapping:
                s_real = self.find_col(source_f2, h2)
                if not s_real:
                    raise Exception(S["ERR_SRC_MISSING"].format(col=source_f2))
                t_real = self.find_col(target_f1, h1)
                if not t_real:
                    raise Exception(S["ERR_DST_MISSING"].format(col=target_f1))
                final_mapping.append((t_real, s_real))

            # Chargement F2 en mémoire (colonnes mappées uniquement)
            self.log(S["LOG_LOADING_F2"])
            cols_utiles = {src for _, src in final_mapping}
            case_sensitive = self.case_sensitive.get()
            f2_dict = {}

            if self._est_excel(tmp_f2):
                for row in self._lire_lignes_excel(tmp_f2, cols_utiles | {k2}):
                    cle = self.nettoyer_cellule(row.get(k2))
                    if not case_sensitive: cle = cle.lower()
                    f2_dict[cle] = {src: row.get(src, "") for src in cols_utiles}
            else:
                enc_f2 = self._detecter_encodage(tmp_f2)
                with open(tmp_f2, 'r', encoding=enc_f2, errors='replace') as f2_data:
                    for row in csv.DictReader(f2_data, delimiter=';'):
                        cle = self.nettoyer_cellule(row.get(k2))
                        if not case_sensitive: cle = cle.lower()
                        f2_dict[cle] = {src: self.nettoyer_cellule(row.get(src, "")) for src in cols_utiles}

            name_part, ext_part = os.path.splitext(out_name)
            self.archiver_si_existe(res_path)
            out_rejet_name = f"{name_part}_Rejet{ext_part}"
            rej_path = os.path.join(out_dir, out_rejet_name)
            if self.create_rejet_file.get():
                self.archiver_si_existe(rej_path)

            nb_entree = 0; nb_ok = 0; nb_rejet = 0; nb_avert = 0
            fmt_sortie = self.out_format.get()

            self.log(S["LOG_START"])

            # Lecture F1 (générateur universel)
            f1_handle = None
            if self._est_excel(tmp_f1):
                source_rows = self._lire_lignes_excel(tmp_f1)
            else:
                enc_f1 = self._detecter_encodage(tmp_f1)
                f1_handle = open(tmp_f1, 'r', encoding=enc_f1, errors='replace')
                source_rows = csv.DictReader(f1_handle, delimiter=';')

            # Parcours F1 — ordre préservé
            lignes_ordonnees = []
            try:
                for row in source_rows:
                    nb_entree += 1
                    row = {k: self.nettoyer_cellule(v) for k, v in row.items() if k is not None}
                    if len(row) != len(h1):
                        nb_avert += 1
                        self.log(S["LOG_LINE_WARN"].format(
                            num=nb_entree, fields=len(row),
                            expected=len(h1), key=row.get(k1, '?')
                        ))
                    cle = self.nettoyer_cellule(row.get(k1))
                    if not case_sensitive: cle = cle.lower()
                    if cle in f2_dict:
                        nb_ok += 1
                        for target, source in final_mapping:
                            row[target] = f2_dict[cle][source]
                        lignes_ordonnees.append((row, True))
                    else:
                        nb_rejet += 1
                        lignes_ordonnees.append((row, False))
            finally:
                if f1_handle: f1_handle.close()

            lignes_ok  = [r for r, ok in lignes_ordonnees if ok]
            lignes_rej = [r for r, ok in lignes_ordonnees if not ok]

            # Écriture du résultat
            if fmt_sortie == "xlsx":
                if self.create_rejet_file.get():
                    lignes_resultat = lignes_ok
                    rej_pour_xlsx   = lignes_rej
                else:
                    lignes_resultat = [r for r, _ in lignes_ordonnees]
                    rej_pour_xlsx   = None
                self._ecrire_xlsx(res_path, h1, lignes_resultat, rej_pour_xlsx)
                suffix = (" " + S["LOG_REJECT_SHEET"].format(count=len(lignes_rej))
                          if self.create_rejet_file.get() else "")
                self.log(S["LOG_RESULT_XLSX"].format(name=out_name) + suffix)
            else:
                with open(res_path, 'w', encoding='utf-8', newline='') as f_out:
                    w = csv.DictWriter(f_out, fieldnames=h1, delimiter=';', extrasaction='ignore')
                    w.writeheader()
                    for row, enrichie in lignes_ordonnees:
                        if enrichie or not self.create_rejet_file.get():
                            w.writerow(row)
                if self.create_rejet_file.get():
                    with open(rej_path, 'w', encoding='utf-8', newline='') as f_rej:
                        w = csv.DictWriter(f_rej, fieldnames=h1, delimiter=';', extrasaction='ignore')
                        w.writeheader()
                        w.writerows(lignes_rej)
                self.convertir_en_ansi(res_path)
                if self.create_rejet_file.get():
                    self.convertir_en_ansi(rej_path)

            pc_ok  = round(nb_ok    / nb_entree * 100) if nb_entree > 0 else 0
            pc_rej = round(nb_rejet / nb_entree * 100) if nb_entree > 0 else 0
            self.log(S["LOG_SEPARATOR"])
            self.log(S["LOG_BILAN"])
            self.log(S["LOG_LINES_READ"].format(count=nb_entree))
            self.log(S["LOG_LINES_OK"].format(count=nb_ok, pct=pc_ok))
            self.log(S["LOG_LINES_REJ"].format(count=nb_rejet, pct=pc_rej))
            if nb_avert > 0:
                self.log(S["LOG_SUSPECT"].format(count=nb_avert))
            self.log(S["LOG_SEPARATOR"])
            if self.create_rejet_file.get() and fmt_sortie != "xlsx":
                self.log(S["LOG_REJET_FILE"].format(name=out_rejet_name))
            if getattr(self, "open_explorer", True):
                os.startfile(os.path.abspath(out_dir))

        except Exception as e:
            self.log(S["LOG_ERROR"].format(msg=str(e)))
        finally:
            for tmp in (tmp_f1, tmp_f2):
                if tmp and tmp not in (f1_path, f2_path) and os.path.exists(tmp):
                    try: os.remove(tmp)
                    except: pass

    # ------------------------------------------------------------------
    # Configuration
    # ------------------------------------------------------------------

    def save_config(self):
        cfg = {
            "lang":           i18n.get_lang(),
            "f1":             self.path_f1.get(),
            "f2":             self.path_f2.get(),
            "result":         self.path_result.get(),
            "out_format":     self.out_format.get(),
            "k1":             self.key_f1.get(),
            "k2":             self.key_f2.get(),
            "rejet":          self.create_rejet_file.get(),
            "case_sensitive": self.case_sensitive.get(),
            "commentaire":    self.comment_area.get("1.0", tk.END),
            "mapping":        self.get_active_mapping()
        }
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=4, ensure_ascii=False)
        self.log(S["LOG_CONFIG_SAVED"])

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    c = json.load(f)
                self.path_f1.set(c.get("f1", ""))
                self.path_f2.set(c.get("f2", ""))
                self.path_result.set(c.get("result", ""))
                self.out_format.set(c.get("out_format", "csv"))
                self.key_f1.set(c.get("k1", "id"))
                self.key_f2.set(c.get("k2", "id"))
                self.create_rejet_file.set(c.get("rejet", False))
                self.case_sensitive.set(c.get("case_sensitive", False))
                self.comment_area.insert("1.0", c.get("commentaire", ""))
                for a, b in c.get("mapping", []):
                    self.add_mapping_row(a, b)
            except Exception:
                self.add_mapping_row("", "")
        else:
            self.add_mapping_row("", "")


if __name__ == "__main__":
    root = tk.Tk()
    app = AppRechercheV(root)
    root.mainloop()
