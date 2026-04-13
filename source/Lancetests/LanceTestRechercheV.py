# -*- coding: utf-8 -*-
"""
TEST SUITE : rechercheV — Suite complète
USAGE :  python LanceTestRechercheV.py
Groupes :
  CAS  — Fonctionnalités CSV de base (14 cas)
  A    — CSV seul (4 cas)
  B    — F1=XLSX / F2=CSV (6 cas)
  C    — F1=CSV / F2=XLSX (5 cas)
  D    — F1=XLSX / F2=XLSX (6 cas)
  E    — Cas limites Excel (4 cas)
  F    — Sortie en XLSX (7 cas)
Les résultats sont conservés dans workdir/
"""

import os, sys, csv, time, shutil, importlib, importlib.util, types, traceback as _tb
from pathlib import Path

WORK_DIR = Path(__file__).parent / "workdir"
SEP = ";"
OK   = "\033[92m✔\033[0m"
KO   = "\033[91m✘\033[0m"
INFO = "\033[94m→\033[0m"

# ══════════════════════════════════════════════════════════════════════════════
# INFRASTRUCTURE
# ══════════════════════════════════════════════════════════════════════════════

def demander_chemin_src() -> Path:
    defaut = Path(__file__).parent.parent / "RechercheV" / "rechercheV.py"
    print("\n" + "═"*62)
    print("  TEST SUITE — rechercheV (suite complète)")
    print("═"*62)
    print(f"  Chemin du programme rechercheV.py")
    print(f"  [Entrée = {defaut}]")
    saisie = input("  > ").strip().strip('"').strip("'")
    p = Path(saisie) if saisie else defaut
    if not p.exists():
        print(f"\n  {KO}  Fichier introuvable : {p}")
        sys.exit(1)
    print(f"  ✔  Fichier trouvé : {p}\n")
    return p

def _mock_tkinter():
    class Var:
        def __init__(self, value=None): self._v = value
        def get(self): return self._v
        def set(self, v): self._v = v

    class W:
        def __init__(self, *a, **kw): pass
        def pack(self, **kw): return self
        def grid(self, **kw): return self
        def bind(self, *a, **kw): pass
        def configure(self, **kw): pass
        def destroy(self): pass
        def after(self, *a, **kw): pass
        def update_idletasks(self): pass
        def insert(self, *a, **kw): pass
        def delete(self, *a, **kw): pass
        def see(self, *a, **kw): pass
        def get(self, *a, **kw): return ""
        def mainloop(self): pass
        def columnconfigure(self, *a, **kw): pass
        def create_window(self, *a, **kw): return 0
        def itemconfig(self, *a, **kw): pass
        def bbox(self, *a, **kw): return (0, 0, 0, 0)

    tk  = types.ModuleType("tkinter")
    ttk = types.ModuleType("tkinter.ttk")
    fd  = types.ModuleType("tkinter.filedialog")
    mb  = types.ModuleType("tkinter.messagebox")
    for m in (tk, ttk):
        for n in ("Tk","Frame","LabelFrame","Label","Entry","Button",
                  "Text","Canvas","Scrollbar","Checkbutton"):
            setattr(m, n, W)
    tk.BooleanVar = Var; tk.StringVar = Var; tk.END = "end"
    tk.TclError = Exception; tk.ttk = ttk; tk.filedialog = fd; tk.messagebox = mb
    sys.modules["tkinter"] = tk; sys.modules["tkinter.ttk"] = ttk
    sys.modules["tkinter.filedialog"] = fd; sys.modules["tkinter.messagebox"] = mb

def _load_source(src_file: Path):
    _mock_tkinter()
    # Ajoute le répertoire source au chemin pour que i18n, strings_fr, strings_en soient importables
    src_dir = str(src_file.parent)
    if src_dir not in sys.path:
        sys.path.insert(0, src_dir)
    spec = importlib.util.spec_from_file_location("rechercheV_src", src_file)
    mod  = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod

def make_app(mod):
    from unittest.mock import MagicMock
    app = object.__new__(mod.AppRechercheV)
    app.root = None; app.mapping_rows = []
    app.log_area = MagicMock(); app.log = lambda t: None
    app.open_explorer = False
    return app

def V(v):
    return type("V", (), {"get": lambda s, _v=v: _v})()

def run_enrichissement(app, mod, f1, f2, out_dir, out_name,
                       k1, k2, mapping_pairs, rejet=False,
                       case_sensitive=False, out_format="csv"):
    app.path_f1           = V(str(f1))
    app.path_f2           = V(str(f2))
    app.path_result       = V(str(Path(out_dir) / out_name))
    app.out_format        = V(out_format)
    app.key_f1            = V(k1)
    app.key_f2            = V(k2)
    app.create_rejet_file = V(rejet)
    app.case_sensitive    = V(case_sensitive)
    app.mapping_rows      = [(V(a), V(b)) for a, b in mapping_pairs]
    app._log_lines = []
    app.log = lambda t: app._log_lines.append(t)
    app.run_process()
    app._run_config = {
        "f1": str(f1), "f2": str(f2), "out_name": out_name,
        "cle_f1": k1, "cle_f2": k2, "mapping": mapping_pairs,
        "rejet": rejet, "case_sensitive": case_sensitive, "out_format": out_format,
    }

# ── Helpers fichiers ──────────────────────────────────────────────────────────

def cas_dir(nom_cas: str) -> Path:
    d = WORK_DIR / nom_cas.replace(" ", "_").replace("—", "-")
    if d.exists(): shutil.rmtree(d)
    d.mkdir(parents=True)
    return d

def write_csv(path, rows, encoding="utf-8-sig"):
    with open(path, "w", newline="", encoding=encoding) as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()), delimiter=SEP)
        w.writeheader(); w.writerows(rows)

def write_xlsx(path, rows):
    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    wb.save(path)

def read_csv(path, encoding="cp1252"):
    with open(path, "r", encoding=encoding, errors="replace") as f:
        return list(csv.DictReader(f, delimiter=SEP))

def read_header(path, encoding="cp1252"):
    with open(path, "r", encoding=encoding, errors="replace") as f:
        return next(csv.reader(f, delimiter=SEP))

def read_xlsx(path, sheet="Résultat"):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True)); wb.close()
    if not rows: return []
    headers = [str(c) if c is not None else "" for c in rows[0]]
    return [dict(zip(headers, [str(v) if v is not None else "" for v in row]))
            for row in rows[1:]]

def xlsx_sheet_exists(path, sheet):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    exists = sheet in wb.sheetnames; wb.close()
    return exists

def count_rows(path: Path):
    """Nombre de lignes de données (hors header) — CSV ou XLSX."""
    if not path.exists(): return "—"
    ext = Path(path).suffix.lower()
    if ext == ".xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
            n = max(0, wb.active.max_row - 1); wb.close(); return n
        except Exception: return "?"
    try:
        with open(path, "r", encoding="cp1252", errors="replace") as f:
            return sum(1 for _ in f) - 1
    except Exception: return "?"

# ── Logging ───────────────────────────────────────────────────────────────────

def write_log(d: Path, nom_cas: str, app, statut: str,
              erreur: str = "", traceback_str: str = ""):
    out_name  = app._run_config["out_name"]
    out_path  = d / out_name
    ext_part  = Path(out_name).suffix.lower()
    rej_path  = d / f"{Path(out_name).stem}_Rejet{ext_part}"

    nb_entree = count_rows(d / Path(app._run_config["f1"]).name)
    nb_sortie = count_rows(out_path) if out_path.exists() else "—"
    nb_rejet  = count_rows(rej_path) if rej_path.exists() else ("0" if app._run_config["rejet"] else "—")

    L = []
    L.append("=" * 62)
    L.append(f"  LOG DE TEST : {nom_cas}")
    L.append(f"  Date       : {time.strftime('%Y-%m-%d %H:%M:%S')}")
    L.append(f"  Résultat   : {'✔ SUCCÈS' if statut == 'OK' else '✘ ÉCHEC'}")
    if erreur: L.append(f"  Erreur     : {erreur}")
    L += ["=" * 62, ""]
    L.append("── CONFIGURATION ──────────────────────────────────────────")
    L.append(f"  Fichier F1         : {app._run_config['f1']}")
    L.append(f"  Fichier F2         : {app._run_config['f2']}")
    L.append(f"  Fichier résultat   : {out_name}")
    L.append(f"  Clé liaison F1     : {app._run_config['cle_f1']}")
    L.append(f"  Clé liaison F2     : {app._run_config['cle_f2']}")
    L.append(f"  Sensible à la casse: {'Oui' if app._run_config['case_sensitive'] else 'Non'}")
    L.append(f"  Fichier de rejet   : {'Oui' if app._run_config['rejet'] else 'Non'}")
    L.append(f"  Format sortie      : {app._run_config.get('out_format','csv').upper()}")
    L.append("")
    L.append("── MAPPING (Destination F1 ← Source F2) ───────────────────")
    for dst, src in app._run_config["mapping"]:
        L.append(f"  {dst:<25} ← {src}")
    L.append("")
    L.append("── COMPTE-RENDU ────────────────────────────────────────────")
    L.append(f"  Lignes lues  (F1)  : {nb_entree}")
    L.append(f"  Lignes en sortie   : {nb_sortie}")
    L.append(f"  Lignes en rejet    : {nb_rejet}")
    L.append("")
    if statut != "OK" and out_path.exists() and ext_part == ".csv":
        try:
            rows = read_csv(out_path)
            L.append("── APERÇU DU RÉSULTAT (5 premières lignes) ────────────────")
            for i, r in enumerate(rows[:5]): L.append(f"  [{i}] {dict(r)}")
            L.append("")
        except Exception: pass
    L.append("── LOG INTERNE ─────────────────────────────────────────────")
    if app._log_lines:
        for l in app._log_lines: L.append(f"  {l}")
    else:
        L.append("  (aucun log interne — le traitement n'a pas démarré)")
    L.append("")
    if traceback_str:
        L.append("── TRACEBACK ───────────────────────────────────────────────")
        for l in traceback_str.strip().splitlines(): L.append(f"  {l}")
        L.append("")
    (d / "test.log").write_text("\n".join(L), encoding="utf-8")

def merge_logs(global_log: Path):
    header = ["═"*62, f"  LOG GLOBAL — {time.strftime('%Y-%m-%d %H:%M:%S')}", "═"*62, ""]
    with open(global_log, "w", encoding="utf-8") as out:
        out.write("\n".join(header) + "\n")
        for lf in sorted(WORK_DIR.rglob("test.log")):
            try: out.write(lf.read_text(encoding="utf-8") + "\n")
            except Exception as e: out.write(f"  [warn] {lf} : {e}\n\n")

TESTS = []
def cas(nom):
    def deco(fn): TESTS.append((nom, fn)); return fn
    return deco

# ══════════════════════════════════════════════════════════════════════════════
# DONNÉES COMMUNES
# ══════════════════════════════════════════════════════════════════════════════

# F1 étendu (14 cas CSV de base — 12 lignes, 8 colonnes)
F1_CLIENTS_FULL = [
    {"client_id":"C001","nom":"Dupont",   "prenom":"Jean",   "ville":"","code_postal":"","contrat":"","segment":"","statut":"Actif"},
    {"client_id":"C002","nom":"Martin",   "prenom":"Sophie", "ville":"","code_postal":"","contrat":"","segment":"","statut":"Actif"},
    {"client_id":"C003","nom":"Bernard",  "prenom":"Paul",   "ville":"","code_postal":"","contrat":"","segment":"","statut":"Inactif"},
    {"client_id":"C004","nom":"Petit",    "prenom":"Marie",  "ville":"","code_postal":"","contrat":"","segment":"","statut":"Actif"},
    {"client_id":"C005","nom":"Robert",   "prenom":"Luc",    "ville":"","code_postal":"","contrat":"","segment":"","statut":"Actif"},
    {"client_id":"C006","nom":"Leroy",    "prenom":"Emma",   "ville":"","code_postal":"","contrat":"","segment":"","statut":"Suspendu"},
    {"client_id":"C007","nom":"Moreau",   "prenom":"Pierre", "ville":"","code_postal":"","contrat":"","segment":"","statut":"Actif"},
    {"client_id":"C008","nom":"Laurent",  "prenom":"Julie",  "ville":"","code_postal":"","contrat":"","segment":"","statut":"Actif"},
    {"client_id":"C009","nom":"Simon",    "prenom":"Marc",   "ville":"","code_postal":"","contrat":"","segment":"","statut":"Inactif"},
    {"client_id":"C010","nom":"Michel",   "prenom":"Claire", "ville":"","code_postal":"","contrat":"","segment":"","statut":"Actif"},
    {"client_id":"ZZZZ","nom":"Inconnu1", "prenom":"X",      "ville":"","code_postal":"","contrat":"","segment":"","statut":"?"},
    {"client_id":"YYYY","nom":"Inconnu2", "prenom":"Y",      "ville":"","code_postal":"","contrat":"","segment":"","statut":"?"},
]

# F2 étendu (pour les 14 cas CSV)
F2_REF_FULL = [
    {"sf_id":"C001","city":"Paris",      "cp":"75001","contrat_type":"PRO","segment_client":"Grand compte","region":"IDF","pays":"France","date_maj":"2024-01-15"},
    {"sf_id":"C002","city":"Lyon",       "cp":"69002","contrat_type":"PME","segment_client":"Moyen compte","region":"ARA","pays":"France","date_maj":"2024-02-10"},
    {"sf_id":"C003","city":"Marseille",  "cp":"13001","contrat_type":"TPE","segment_client":"Petit compte","region":"PACA","pays":"France","date_maj":"2024-01-20"},
    {"sf_id":"C004","city":"Bordeaux",   "cp":"33000","contrat_type":"PRO","segment_client":"Grand compte","region":"NAQ","pays":"France","date_maj":"2024-03-05"},
    {"sf_id":"C005","city":"Nantes",     "cp":"44000","contrat_type":"PME","segment_client":"Moyen compte","region":"PDL","pays":"France","date_maj":"2024-02-28"},
    {"sf_id":"C006","city":"Strasbourg", "cp":"67000","contrat_type":"TPE","segment_client":"Petit compte","region":"GES","pays":"France","date_maj":"2024-01-08"},
    {"sf_id":"C007","city":"Toulouse",   "cp":"31000","contrat_type":"PRO","segment_client":"Grand compte","region":"OCC","pays":"France","date_maj":"2024-03-12"},
    {"sf_id":"C008","city":"Nice",       "cp":"06000","contrat_type":"PME","segment_client":"Moyen compte","region":"PACA","pays":"France","date_maj":"2024-02-14"},
    {"sf_id":"c009","city":"Rennes",     "cp":"35000","contrat_type":"TPE","segment_client":"Petit compte","region":"BRE","pays":"France","date_maj":"2024-01-30"},  # casse minuscule
    {"sf_id":"C010","city":"Lille",      "cp":"59000","contrat_type":"PRO","segment_client":"Grand compte","region":"HDF","pays":"France","date_maj":"2024-03-20"},
    {"sf_id":"C011","city":"Grenoble",   "cp":"38000","contrat_type":"PME","segment_client":"Moyen compte","region":"ARA","pays":"France","date_maj":"2024-02-05"},  # orphelin
]

# F1 réduit (groupes A-F — 6 lignes)
F1_CLIENTS = [
    {"client_id":"C001","nom":"Dupont",  "prenom":"Jean",   "ville":"","code_postal":"","contrat":"","segment":"","statut":"Actif"},
    {"client_id":"C002","nom":"Martin",  "prenom":"Sophie", "ville":"","code_postal":"","contrat":"","segment":"","statut":"Actif"},
    {"client_id":"C003","nom":"Bernard", "prenom":"Paul",   "ville":"","code_postal":"","contrat":"","segment":"","statut":"Inactif"},
    {"client_id":"C004","nom":"Petit",   "prenom":"Marie",  "ville":"","code_postal":"","contrat":"","segment":"","statut":"Actif"},
    {"client_id":"C005","nom":"Robert",  "prenom":"Luc",    "ville":"","code_postal":"","contrat":"","segment":"","statut":"Actif"},
    {"client_id":"ZZZZ","nom":"Inconnu", "prenom":"X",      "ville":"","code_postal":"","contrat":"","segment":"","statut":"?"},
]

# F2 réduit (groupes A-F)
F2_REF = [
    {"sf_id":"C001","city":"Paris",     "cp":"75001","contrat_type":"PRO","segment_client":"Grand compte"},
    {"sf_id":"C002","city":"Lyon",      "cp":"69002","contrat_type":"PME","segment_client":"Moyen compte"},
    {"sf_id":"C003","city":"Marseille", "cp":"13001","contrat_type":"TPE","segment_client":"Petit compte"},
    {"sf_id":"C004","city":"Bordeaux",  "cp":"33000","contrat_type":"PRO","segment_client":"Grand compte"},
    {"sf_id":"C005","city":"Nantes",    "cp":"44000","contrat_type":"PME","segment_client":"Moyen compte"},
    {"sf_id":"C006","city":"Grenoble",  "cp":"38000","contrat_type":"TPE","segment_client":"Petit compte"},  # orphelin
]

MAPPING_BASE = [("ville","city"),("code_postal","cp"),("contrat","contrat_type")]

def assert_enrichissement_ok(r):
    """Vérifie les résultats communs aux groupes A-F (6 lignes, 5 enrichies)."""
    assert len(r) == 6,                    f"Attendu 6 lignes, obtenu {len(r)}"
    assert r[0]["ville"] == "Paris",       f"C001 ville attendu Paris, obtenu '{r[0]['ville']}'"
    assert r[1]["ville"] == "Lyon",        f"C002 ville attendu Lyon, obtenu '{r[1]['ville']}'"
    assert r[2]["code_postal"] == "13001", f"C003 cp attendu 13001, obtenu '{r[2]['code_postal']}'"
    assert r[3]["contrat"] == "PRO",       f"C004 contrat attendu PRO, obtenu '{r[3]['contrat']}'"
    assert r[4]["ville"] == "Nantes",      f"C005 ville attendu Nantes, obtenu '{r[4]['ville']}'"
    assert r[5]["ville"] == "",            f"ZZZZ doit rester vide, obtenu '{r[5]['ville']}'"

# ══════════════════════════════════════════════════════════════════════════════
# GROUPE CAS — FONCTIONNALITÉS CSV DE BASE (14 cas)
# ══════════════════════════════════════════════════════════════════════════════

@cas("CAS01-enrichissement_standard")
def t01(app, mod):
    """Enrichissement standard : 10 clients trouvés, 2 sans correspondance (ZZZZ, YYYY).
    Attendu : 12 lignes dans le résultat, colonnes enrichies pour C001-C010."""
    d = cas_dir("CAS01-enrichissement_standard")
    f1 = d/"f1_clients.csv"; f2 = d/"f2_ref.csv"
    write_csv(f1, F1_CLIENTS_FULL); write_csv(f2, F2_REF_FULL)
    run_enrichissement(app, mod, f1, f2, d, "out_resultat.csv", "client_id", "sf_id",
                       [("ville","city"),("code_postal","cp"),("contrat","contrat_type")])
    r = read_csv(d/"out_resultat.csv")
    assert len(r) == 12,             f"Attendu 12 lignes, obtenu {len(r)}"
    assert r[0]["ville"] == "Paris", f"C001 attendu Paris, obtenu '{r[0]['ville']}'"
    assert r[1]["ville"] == "Lyon",  f"C002 attendu Lyon, obtenu '{r[1]['ville']}'"
    assert r[3]["code_postal"] == "33000", f"C004 cp attendu 33000, obtenu '{r[3]['code_postal']}'"
    assert r[4]["contrat"] == "PME", f"C005 contrat attendu PME, obtenu '{r[4]['contrat']}'"
    assert r[10]["ville"] == "",     f"ZZZZ doit rester vide, obtenu '{r[10]['ville']}'"
    assert r[11]["ville"] == "",     f"YYYY doit rester vide, obtenu '{r[11]['ville']}'"

@cas("CAS02-fichier_rejet")
def t02(app, mod):
    """Lignes sans correspondance (ZZZZ, YYYY) dans le fichier de rejet.
    Attendu : résultat=10, rejet=2 (ZZZZ+YYYY). C011 orphelin dans F2 ignoré."""
    d = cas_dir("CAS02-fichier_rejet")
    f1 = d/"f1_clients.csv"; f2 = d/"f2_ref.csv"
    write_csv(f1, F1_CLIENTS_FULL); write_csv(f2, F2_REF_FULL)
    run_enrichissement(app, mod, f1, f2, d, "out_resultat.csv", "client_id", "sf_id",
                       [("ville","city"),("code_postal","cp")], rejet=True)
    ok  = read_csv(d/"out_resultat.csv")
    rej = read_csv(d/"out_resultat_Rejet.csv")
    assert len(ok) == 10,  f"Enrichi : attendu 10, obtenu {len(ok)}"
    assert len(rej) == 2,  f"Rejet : attendu 2, obtenu {len(rej)}"
    assert {r["client_id"] for r in rej} == {"ZZZZ","YYYY"}

@cas("CAS03-insensible_casse")
def t03(app, mod):
    """C009 majuscule dans F1 doit matcher c009 minuscule dans F2.
    Attendu : ville = Rennes."""
    d = cas_dir("CAS03-insensible_casse")
    f1 = d/"f1_clients.csv"; f2 = d/"f2_ref.csv"
    write_csv(f1, F1_CLIENTS_FULL); write_csv(f2, F2_REF_FULL)
    run_enrichissement(app, mod, f1, f2, d, "out_resultat.csv", "client_id", "sf_id",
                       [("ville","city"),("code_postal","cp")], case_sensitive=False)
    r = read_csv(d/"out_resultat.csv")
    assert r[8]["ville"] == "Rennes", f"C009 attendu Rennes, obtenu '{r[8]['ville']}'"

@cas("CAS04-sensible_casse")
def t04(app, mod):
    """C009 majuscule ne doit PAS matcher c009 minuscule en mode sensible.
    Attendu : ville = ''."""
    d = cas_dir("CAS04-sensible_casse")
    f1 = d/"f1_clients.csv"; f2 = d/"f2_ref.csv"
    write_csv(f1, F1_CLIENTS_FULL); write_csv(f2, F2_REF_FULL)
    run_enrichissement(app, mod, f1, f2, d, "out_resultat.csv", "client_id", "sf_id",
                       [("ville","city"),("code_postal","cp")], case_sensitive=True)
    r = read_csv(d/"out_resultat.csv")
    assert r[8]["ville"] == "", f"C009 sensible casse, attendu vide, obtenu '{r[8]['ville']}'"

@cas("CAS05-trim_espaces_cle")
def t05(app, mod):
    """Clés avec espaces parasites en début/fin dans F1 matchent après strip().
    Attendu : les 4 lignes enrichies malgré les espaces."""
    d = cas_dir("CAS05-trim_espaces_cle")
    f1 = d/"f1_espaces.csv"; f2 = d/"f2_ref.csv"
    write_csv(f1, [
        {"ref":"  PCE001  ","adresse":"","ville":"","cp":"","tarif":"","commentaire":""},
        {"ref":" PCE002",   "adresse":"","ville":"","cp":"","tarif":"","commentaire":""},
        {"ref":"PCE003 ",   "adresse":"","ville":"","cp":"","tarif":"","commentaire":""},
        {"ref":"PCE004",    "adresse":"","ville":"","cp":"","tarif":"","commentaire":""},
    ])
    write_csv(f2, [
        {"pce":"PCE001","adr":"12 rue de la Paix",  "cit":"Paris",    "codepost":"75001","prix":"0.15"},
        {"pce":"PCE002","adr":"5 avenue Foch",       "cit":"Lyon",     "codepost":"69002","prix":"0.12"},
        {"pce":"PCE003","adr":"8 bd du Port",        "cit":"Marseille","codepost":"13001","prix":"0.18"},
        {"pce":"PCE004","adr":"3 rue Neuve",          "cit":"Bordeaux", "codepost":"33000","prix":"0.10"},
    ])
    run_enrichissement(app, mod, f1, f2, d, "out_resultat.csv", "ref", "pce",
                       [("adresse","adr"),("ville","cit"),("cp","codepost"),("tarif","prix")])
    r = read_csv(d/"out_resultat.csv")
    assert r[0]["ville"] == "Paris",     f"PCE001 attendu Paris, obtenu '{r[0]['ville']}'"
    assert r[1]["ville"] == "Lyon",      f"PCE002 attendu Lyon, obtenu '{r[1]['ville']}'"
    assert r[2]["ville"] == "Marseille", f"PCE003 attendu Marseille, obtenu '{r[2]['ville']}'"
    assert r[3]["ville"] == "Bordeaux",  f"PCE004 attendu Bordeaux, obtenu '{r[3]['ville']}'"

@cas("CAS06-espaces_internes_preserves")
def t06(app, mod):
    """Espaces multiples internes aux valeurs enrichies non normalisés.
    Attendu : doubles/triples espaces conservés tels quels."""
    d = cas_dir("CAS06-espaces_internes_preserves")
    f1 = d/"f1.csv"; f2 = d/"f2.csv"
    write_csv(f1, [{"id":"R01","libelle":"","adresse":"","commentaire":""},
                   {"id":"R02","libelle":"","adresse":"","commentaire":""},
                   {"id":"R03","libelle":"","adresse":"","commentaire":""}])
    write_csv(f2, [
        {"id":"R01","lib":"Saint  Germain  en  Laye","adr":"12  rue  du  Château",            "com":"Ref  A"},
        {"id":"R02","lib":"Le   Mans",               "adr":"5   place   de   la   République", "com":""},
        {"id":"R03","lib":"Aix en Provence",          "adr":"Route  Nationale  7",             "com":"Double  vérif"},
    ])
    run_enrichissement(app, mod, f1, f2, d, "out_resultat.csv", "id", "id",
                       [("libelle","lib"),("adresse","adr"),("commentaire","com")])
    r = read_csv(d/"out_resultat.csv")
    assert r[0]["libelle"] == "Saint  Germain  en  Laye",        f"Espaces altérés : '{r[0]['libelle']}'"
    assert r[1]["adresse"] == "5   place   de   la   République", f"Espaces altérés : '{r[1]['adresse']}'"

@cas("CAS07-mapping_multiple_5_colonnes")
def t07(app, mod):
    """Mapping de 5 colonnes simultanément sur 15 lignes.
    Attendu : toutes colonnes mappées enrichies, colonnes inutiles absentes."""
    d = cas_dir("CAS07-mapping_multiple_5_colonnes")
    f1 = d/"f1_pce.csv"; f2 = d/"f2_enrichissement.csv"
    f1_rows = [{"pce_ref":f"PCE{i:03d}","nom_site":"","adresse":"","cp":"","ville":"","contrat":"","date_pose":"","segment":""} for i in range(1,16)]
    f2_rows = [{"ref":f"PCE{i:03d}","site":f"Site_{i:03d}","adr":f"{i} rue de la Liberté",
                "code":f"{10000+i*100}","cit":f"Ville_{i:03d}","ctr":"PRO" if i%2==0 else "PME",
                "colonne_inutile_A":"xxx","colonne_inutile_B":"yyy"} for i in range(1,16)]
    write_csv(f1, f1_rows); write_csv(f2, f2_rows)
    run_enrichissement(app, mod, f1, f2, d, "out_resultat.csv", "pce_ref", "ref",
                       [("nom_site","site"),("adresse","adr"),("cp","code"),("ville","cit"),("contrat","ctr")])
    r = read_csv(d/"out_resultat.csv")
    assert len(r) == 15,                   f"Attendu 15 lignes, obtenu {len(r)}"
    assert r[0]["nom_site"] == "Site_001", f"Site_001 attendu, obtenu '{r[0]['nom_site']}'"
    assert r[7]["contrat"] == "PRO",       f"PCE008 contrat PRO attendu, obtenu '{r[7]['contrat']}'"
    assert "colonne_inutile_A" not in r[0], "colonne_inutile_A ne doit pas apparaître"

@cas("CAS08-f1_vide")
def t08(app, mod):
    """F1 avec uniquement le header (0 ligne de données).
    Attendu : résultat créé avec 0 ligne, pas de crash."""
    d = cas_dir("CAS08-f1_vide")
    f1 = d/"f1_vide.csv"; f2 = d/"f2_ref.csv"
    with open(f1, "w", encoding="utf-8-sig") as fh:
        fh.write("client_id;nom;prenom;ville;code_postal;contrat;segment;statut\n")
    write_csv(f2, F2_REF_FULL)
    run_enrichissement(app, mod, f1, f2, d, "out_resultat.csv", "client_id", "sf_id",
                       [("ville","city"),("code_postal","cp")])
    assert len(read_csv(d/"out_resultat.csv")) == 0

@cas("CAS09-encodage_cp1252")
def t09(app, mod):
    """F1 encodé CP1252 avec caractères accentués — auto-converti en UTF-8.
    Attendu : enrichissement correct malgré l'encodage source."""
    d = cas_dir("CAS09-encodage_cp1252")
    f1 = d/"f1_cp1252.csv"; f2 = d/"f2_utf8.csv"
    with open(f1, "w", encoding="cp1252") as fh:
        fh.write("ref;nom;ville;cp;contrat\n")
        fh.write("R001;Dup\xe9ont;;;\n")
        fh.write("R002;Lef\xe8vre;;;\n")
        fh.write("R003;Mich\xe9l;;;\n")
    write_csv(f2, [
        {"ref":"R001","cit":"Paris",    "code":"75001","ctr":"PRO"},
        {"ref":"R002","cit":"Lyon",     "code":"69002","ctr":"PME"},
        {"ref":"R003","cit":"Bordeaux", "code":"33000","ctr":"TPE"},
    ])
    run_enrichissement(app, mod, f1, f2, d, "out_resultat.csv", "ref", "ref",
                       [("ville","cit"),("cp","code"),("contrat","ctr")])
    r = read_csv(d/"out_resultat.csv")
    assert r[0]["ville"] == "Paris",   f"R001 attendu Paris, obtenu '{r[0]['ville']}'"
    assert r[1]["ville"] == "Lyon",    f"R002 attendu Lyon, obtenu '{r[1]['ville']}'"
    assert r[2]["contrat"] == "TPE",   f"R003 contrat attendu TPE, obtenu '{r[2]['contrat']}'"

@cas("CAS10-archivage_fichier_existant")
def t10(app, mod):
    """Fichier résultat préexistant archivé (horodaté) avant réécriture.
    Attendu : ancien archivé avec timestamp, nouveau créé avec données fraîches."""
    d = cas_dir("CAS10-archivage_fichier_existant")
    f1 = d/"f1_clients.csv"; f2 = d/"f2_ref.csv"; out = d/"out_resultat.csv"
    write_csv(f1, F1_CLIENTS_FULL[:5]); write_csv(f2, F2_REF_FULL)
    write_csv(out, F1_CLIENTS_FULL[:1])
    rows_old = read_csv(out, encoding="utf-8-sig"); rows_old[0]["ville"] = "ANCIENNE"
    write_csv(out, rows_old, encoding="utf-8-sig")
    time.sleep(1.1)
    run_enrichissement(app, mod, f1, f2, d, "out_resultat.csv", "client_id", "sf_id",
                       [("ville","city"),("code_postal","cp")])
    archives = list(d.glob("out_resultat_2*.csv"))
    assert len(archives) >= 1, "Aucun fichier archivé trouvé"
    assert read_csv(archives[0])[0]["ville"] == "ANCIENNE"
    assert read_csv(d/"out_resultat.csv")[0]["ville"] == "Paris"

@cas("CAS11-cle_absente_erreur_propre")
def t11(app, mod):
    """Clé inexistante dans F1 → ERREUR claire sans crash.
    Attendu : log contient 'ERREUR'."""
    d = cas_dir("CAS11-cle_absente_erreur_propre")
    f1 = d/"f1_clients.csv"; f2 = d/"f2_ref.csv"
    write_csv(f1, F1_CLIENTS_FULL); write_csv(f2, F2_REF_FULL)
    run_enrichissement(app, mod, f1, f2, d, "out_resultat.csv", "colonne_inexistante", "sf_id",
                       [("ville","city")])
    assert any("ERREUR" in l for l in app._log_lines), "Aucun message ERREUR produit"

@cas("CAS12-chargement_selectif_colonnes_f2")
def t12(app, mod):
    """Seules les colonnes mappées de F2 apparaissent dans le résultat.
    Attendu : region, pays, date_maj absents."""
    d = cas_dir("CAS12-chargement_selectif_colonnes_f2")
    f1 = d/"f1_clients.csv"; f2 = d/"f2_ref.csv"
    write_csv(f1, F1_CLIENTS_FULL); write_csv(f2, F2_REF_FULL)
    run_enrichissement(app, mod, f1, f2, d, "out_resultat.csv", "client_id", "sf_id",
                       [("ville","city"),("segment","segment_client")])
    r = read_csv(d/"out_resultat.csv")
    assert r[0]["ville"] == "Paris",          f"C001 attendu Paris, obtenu '{r[0]['ville']}'"
    assert r[0]["segment"] == "Grand compte", f"C001 segment attendu 'Grand compte', obtenu '{r[0]['segment']}'"
    assert "region"   not in r[0], "La colonne 'region' ne doit pas apparaître"
    assert "pays"     not in r[0], "La colonne 'pays' ne doit pas apparaître"
    assert "date_maj" not in r[0], "La colonne 'date_maj' ne doit pas apparaître"

@cas("CAS13-structure_colonnes_identiques_f1")
def t13(app, mod):
    """Le résultat a EXACTEMENT les mêmes colonnes que F1, dans le même ordre.
    Attendu : header résultat == header F1."""
    d = cas_dir("CAS13-structure_colonnes_identiques_f1")
    f1 = d/"f1_clients.csv"; f2 = d/"f2_ref.csv"
    write_csv(f1, F1_CLIENTS_FULL); write_csv(f2, F2_REF_FULL)
    run_enrichissement(app, mod, f1, f2, d, "out_resultat.csv", "client_id", "sf_id",
                       [("ville","city"),("code_postal","cp"),("contrat","contrat_type"),("segment","segment_client")])
    h_f1  = read_header(d/"f1_clients.csv", encoding="utf-8-sig")
    h_out = read_header(d/"out_resultat.csv")
    assert h_out == h_f1, f"Colonnes résultat ≠ F1\n  F1  : {h_f1}\n  OUT : {h_out}"

@cas("CAS14-retours_chariot_embarques")
def t14(app, mod):
    """Retours chariot \\r\\n embarqués dans les cellules — réparés avant parsing.
    Attendu : 4 lignes enrichies, 0 rejet, log signale le nettoyage."""
    d = cas_dir("CAS14-retours_chariot_embarques")
    f1 = d/"f1_rc.csv"; f2 = d/"f2_rc.csv"
    with open(f1, "wb") as fh:
        fh.write(b"\xef\xbb\xbf")
        fh.write(b"ref;nom;ville;contrat\r\n")
        fh.write(b"A001;Retour\r\nChariot;;\r\n")
        fh.write(b"A002;Martin;;\r\n")
        fh.write(b"A003;Dupont\r\nPaul;;\r\n")
        fh.write(b"A004;Petit;;\r\n")
    with open(f2, "wb") as fh:
        fh.write(b"\xef\xbb\xbf")
        fh.write(b"ref;city;ctr\r\n")
        fh.write(b"A001;Paris;PRO\r\n"); fh.write(b"A002;Lyon;PME\r\n")
        fh.write(b"A003;Marseille;TPE\r\n"); fh.write(b"A004;Bordeaux;PRO\r\n")
    run_enrichissement(app, mod, f1, f2, d, "out_resultat.csv", "ref", "ref",
                       [("ville","city"),("contrat","ctr")], rejet=True)
    r   = read_csv(d/"out_resultat.csv")
    rej = read_csv(d/"out_resultat_Rejet.csv")
    assert not [l for l in app._log_lines if "ERREUR" in l], "Erreur inattendue"
    assert len(r) == 4,   f"Résultat attendu 4, obtenu {len(r)}"
    assert len(rej) == 0, f"Rejet attendu 0, obtenu {len(rej)}"
    assert r[0]["ville"] != "", "A001 ville vide après nettoyage RC"
    assert r[1]["ville"] == "Lyon",      f"A002 attendu Lyon, obtenu '{r[1]['ville']}'"
    assert r[2]["ville"] == "Marseille", f"A003 attendu Marseille, obtenu '{r[2]['ville']}'"
    assert any("nettoyé" in l.lower() or "retour" in l.lower() for l in app._log_lines), \
        "Le log devrait signaler le nettoyage RC"

# ══════════════════════════════════════════════════════════════════════════════
# GROUPE A — CSV SEUL (référence)
# ══════════════════════════════════════════════════════════════════════════════

@cas("A01-CSV-F1_CSV-F2_standard")
def a01(app, mod):
    """F1=CSV, F2=CSV — enrichissement standard de référence.
    Attendu : 5 lignes enrichies, ZZZZ reste vide."""
    d = cas_dir("A01-CSV-F1_CSV-F2_standard")
    f1 = d/"f1.csv"; f2 = d/"f2.csv"
    write_csv(f1, F1_CLIENTS); write_csv(f2, F2_REF)
    run_enrichissement(app, mod, f1, f2, d, "out.csv", "client_id", "sf_id", MAPPING_BASE)
    assert_enrichissement_ok(read_csv(d/"out.csv"))

@cas("A02-CSV-F1_CSV-F2_rejet")
def a02(app, mod):
    """F1=CSV, F2=CSV — avec fichier de rejet.
    Attendu : résultat=5, rejet=1 (ZZZZ)."""
    d = cas_dir("A02-CSV-F1_CSV-F2_rejet")
    f1 = d/"f1.csv"; f2 = d/"f2.csv"
    write_csv(f1, F1_CLIENTS); write_csv(f2, F2_REF)
    run_enrichissement(app, mod, f1, f2, d, "out.csv", "client_id", "sf_id", MAPPING_BASE, rejet=True)
    ok = read_csv(d/"out.csv"); rej = read_csv(d/"out_Rejet.csv")
    assert len(ok) == 5,  f"Enrichi attendu 5, obtenu {len(ok)}"
    assert len(rej) == 1, f"Rejet attendu 1, obtenu {len(rej)}"
    assert rej[0]["client_id"] == "ZZZZ"

@cas("A03-CSV-F1_CSV-F2_insensible_casse")
def a03(app, mod):
    """F1=CSV, F2=CSV — clé en minuscule dans F2, mode insensible.
    Attendu : match malgré la casse."""
    d = cas_dir("A03-CSV-F1_CSV-F2_insensible_casse")
    f2_casse = [dict(r, **{"sf_id": r["sf_id"].lower()}) for r in F2_REF]
    f1 = d/"f1.csv"; f2 = d/"f2.csv"
    write_csv(f1, F1_CLIENTS); write_csv(f2, f2_casse)
    run_enrichissement(app, mod, f1, f2, d, "out.csv", "client_id", "sf_id", MAPPING_BASE, case_sensitive=False)
    assert read_csv(d/"out.csv")[0]["ville"] == "Paris"

@cas("A04-CSV-F1_CSV-F2_encodage_cp1252")
def a04(app, mod):
    """F1=CSV encodé CP1252 — lecture correcte des caractères accentués.
    Attendu : enrichissement correct."""
    d = cas_dir("A04-CSV-F1_CSV-F2_encodage_cp1252")
    f1 = d/"f1_cp1252.csv"; f2 = d/"f2.csv"
    with open(f1, "w", encoding="cp1252") as fh:
        fh.write("client_id;nom;ville;contrat\n")
        fh.write("C001;Dup\xe9ont;;;\n")
        fh.write("C002;L\xe9vy;;;\n")
    write_csv(f2, F2_REF)
    run_enrichissement(app, mod, f1, f2, d, "out.csv", "client_id", "sf_id",
                       [("ville","city"),("contrat","contrat_type")])
    r = read_csv(d/"out.csv")
    assert r[0]["ville"] == "Paris", f"C001 attendu Paris, obtenu '{r[0]['ville']}'"
    assert r[1]["ville"] == "Lyon",  f"C002 attendu Lyon, obtenu '{r[1]['ville']}'"

# ══════════════════════════════════════════════════════════════════════════════
# GROUPE B — F1=XLSX, F2=CSV
# ══════════════════════════════════════════════════════════════════════════════

@cas("B01-XLSX-F1_xlsx_F2_csv_standard")
def b01(app, mod):
    """F1=XLSX, F2=CSV — enrichissement standard. F1 non modifié."""
    d = cas_dir("B01-XLSX-F1_xlsx_F2_csv_standard")
    f1 = d/"f1.xlsx"; f2 = d/"f2.csv"
    write_xlsx(f1, F1_CLIENTS); write_csv(f2, F2_REF)
    mtime = os.path.getmtime(f1)
    run_enrichissement(app, mod, f1, f2, d, "out.csv", "client_id", "sf_id", MAPPING_BASE)
    assert os.path.getmtime(f1) == mtime, "F1 xlsx modifié — INTERDIT"
    assert_enrichissement_ok(read_csv(d/"out.csv"))

@cas("B02-XLSX-F1_xlsx_F2_csv_rejet")
def b02(app, mod):
    """F1=XLSX, F2=CSV — avec fichier de rejet.
    Attendu : résultat=5, rejet=1."""
    d = cas_dir("B02-XLSX-F1_xlsx_F2_csv_rejet")
    f1 = d/"f1.xlsx"; f2 = d/"f2.csv"
    write_xlsx(f1, F1_CLIENTS); write_csv(f2, F2_REF)
    run_enrichissement(app, mod, f1, f2, d, "out.csv", "client_id", "sf_id", MAPPING_BASE, rejet=True)
    ok = read_csv(d/"out.csv"); rej = read_csv(d/"out_Rejet.csv")
    assert len(ok) == 5, f"Enrichi attendu 5, obtenu {len(ok)}"
    assert len(rej) == 1, f"Rejet attendu 1, obtenu {len(rej)}"

@cas("B03-XLSX-F1_xlsx_F2_csv_insensible_casse")
def b03(app, mod):
    """F1=XLSX, F2=CSV — clé minuscule dans F2, mode insensible."""
    d = cas_dir("B03-XLSX-F1_xlsx_F2_csv_insensible_casse")
    f2_casse = [dict(r, **{"sf_id": r["sf_id"].lower()}) for r in F2_REF]
    f1 = d/"f1.xlsx"; f2 = d/"f2.csv"
    write_xlsx(f1, F1_CLIENTS); write_csv(f2, f2_casse)
    run_enrichissement(app, mod, f1, f2, d, "out.csv", "client_id", "sf_id", MAPPING_BASE, case_sensitive=False)
    assert read_csv(d/"out.csv")[0]["ville"] == "Paris"

@cas("B04-XLSX-F1_xlsx_nombres_comme_cle")
def b04(app, mod):
    """F1=XLSX — clé numérique (entier) dans Excel → convertie en str.
    Attendu : '75001' sans '.0'."""
    d = cas_dir("B04-XLSX-F1_xlsx_nombres_comme_cle")
    f1 = d/"f1.xlsx"; f2 = d/"f2.csv"
    write_xlsx(f1, [{"code":75001,"nom":"Paris Centre","region":"","actif":""},
                    {"code":69002,"nom":"Lyon Part-Dieu","region":"","actif":""},
                    {"code":99999,"nom":"Inconnu","region":"","actif":""}])
    write_csv(f2, [{"cp":"75001","reg":"IDF","ok":"1"},{"cp":"69002","reg":"ARA","ok":"1"}])
    run_enrichissement(app, mod, f1, f2, d, "out.csv", "code", "cp", [("region","reg"),("actif","ok")])
    r = read_csv(d/"out.csv")
    assert len(r) == 3,                    f"Attendu 3 lignes, obtenu {len(r)}"
    assert r[0]["region"] == "IDF",        f"75001 région attendue IDF, obtenu '{r[0]['region']}'"
    assert r[0]["code"] == "75001",        f"75001 attendu sans décimale, obtenu '{r[0]['code']}'"
    assert "." not in r[0]["code"],        f"Pas de décimale : '{r[0]['code']}'"
    assert r[2]["region"] == "",           f"99999 doit rester vide"

@cas("B05-XLSX-F1_xlsx_vide")
def b05(app, mod):
    """F1=XLSX vide (header seul). Attendu : 0 ligne de données, pas de crash."""
    d = cas_dir("B05-XLSX-F1_xlsx_vide")
    import openpyxl
    wb = openpyxl.Workbook(); wb.active.append(["client_id","nom","ville","contrat"])
    f1 = d/"f1_vide.xlsx"; wb.save(f1)
    write_csv(d/"f2.csv", F2_REF)
    run_enrichissement(app, mod, f1, d/"f2.csv", d, "out.csv", "client_id", "sf_id",
                       [("ville","city"),("contrat","contrat_type")])
    assert len(read_csv(d/"out.csv")) == 0

@cas("B06-XLSX-F1_xlsx_non_modifie_apres_erreur")
def b06(app, mod):
    """F1=XLSX — erreur (clé absente) : F1 intact même en cas d'erreur."""
    d = cas_dir("B06-XLSX-F1_xlsx_non_modifie_apres_erreur")
    f1 = d/"f1.xlsx"; f2 = d/"f2.csv"
    write_xlsx(f1, F1_CLIENTS); write_csv(f2, F2_REF)
    taille = os.path.getsize(f1); mtime = os.path.getmtime(f1)
    run_enrichissement(app, mod, f1, f2, d, "out.csv", "cle_inexistante", "sf_id", MAPPING_BASE)
    assert any("ERREUR" in l for l in app._log_lines), "Aucun message ERREUR"
    assert os.path.getsize(f1) == taille,  "Taille F1 modifiée — INTERDIT"
    assert os.path.getmtime(f1) == mtime,  "Date F1 modifiée — INTERDIT"

# ══════════════════════════════════════════════════════════════════════════════
# GROUPE C — F1=CSV, F2=XLSX
# ══════════════════════════════════════════════════════════════════════════════

@cas("C01-XLSX-F1_csv_F2_xlsx_standard")
def c01(app, mod):
    """F1=CSV, F2=XLSX — enrichissement standard. F2 non modifié."""
    d = cas_dir("C01-XLSX-F1_csv_F2_xlsx_standard")
    f1 = d/"f1.csv"; f2 = d/"f2.xlsx"
    write_csv(f1, F1_CLIENTS); write_xlsx(f2, F2_REF)
    mtime = os.path.getmtime(f2)
    run_enrichissement(app, mod, f1, f2, d, "out.csv", "client_id", "sf_id", MAPPING_BASE)
    assert os.path.getmtime(f2) == mtime, "F2 xlsx modifié — INTERDIT"
    assert_enrichissement_ok(read_csv(d/"out.csv"))

@cas("C02-XLSX-F1_csv_F2_xlsx_rejet")
def c02(app, mod):
    """F1=CSV, F2=XLSX — avec fichier de rejet. Attendu : résultat=5, rejet=1."""
    d = cas_dir("C02-XLSX-F1_csv_F2_xlsx_rejet")
    f1 = d/"f1.csv"; f2 = d/"f2.xlsx"
    write_csv(f1, F1_CLIENTS); write_xlsx(f2, F2_REF)
    run_enrichissement(app, mod, f1, f2, d, "out.csv", "client_id", "sf_id", MAPPING_BASE, rejet=True)
    ok = read_csv(d/"out.csv"); rej = read_csv(d/"out_Rejet.csv")
    assert len(ok) == 5, f"Enrichi attendu 5, obtenu {len(ok)}"
    assert len(rej) == 1, f"Rejet attendu 1, obtenu {len(rej)}"

@cas("C03-XLSX-F1_csv_F2_xlsx_insensible_casse")
def c03(app, mod):
    """F1=CSV, F2=XLSX — clé minuscule dans F2, mode insensible."""
    d = cas_dir("C03-XLSX-F1_csv_F2_xlsx_insensible_casse")
    f2_casse = [dict(r, **{"sf_id": r["sf_id"].lower()}) for r in F2_REF]
    f1 = d/"f1.csv"; f2 = d/"f2.xlsx"
    write_csv(f1, F1_CLIENTS); write_xlsx(f2, f2_casse)
    run_enrichissement(app, mod, f1, f2, d, "out.csv", "client_id", "sf_id", MAPPING_BASE, case_sensitive=False)
    assert read_csv(d/"out.csv")[0]["ville"] == "Paris"

@cas("C04-XLSX-F1_csv_F2_xlsx_colonnes_inutiles_ignorees")
def c04(app, mod):
    """F1=CSV, F2=XLSX — colonnes non mappées de F2 absentes du résultat."""
    d = cas_dir("C04-XLSX-F1_csv_F2_xlsx_colonnes_inutiles_ignorees")
    f2_extra = [dict(r, **{"col_A":"xxx","col_B":"yyy"}) for r in F2_REF]
    f1 = d/"f1.csv"; f2 = d/"f2.xlsx"
    write_csv(f1, F1_CLIENTS); write_xlsx(f2, f2_extra)
    run_enrichissement(app, mod, f1, f2, d, "out.csv", "client_id", "sf_id",
                       [("ville","city"),("segment","segment_client")])
    r = read_csv(d/"out.csv")
    assert r[0]["ville"] == "Paris", f"C001 attendu Paris, obtenu '{r[0]['ville']}'"
    assert "col_A" not in r[0], "col_A ne doit pas apparaître"

@cas("C05-XLSX-F1_csv_F2_xlsx_non_modifie_apres_erreur")
def c05(app, mod):
    """F1=CSV, F2=XLSX — erreur (clé absente) : F2 intact."""
    d = cas_dir("C05-XLSX-F1_csv_F2_xlsx_non_modifie_apres_erreur")
    f1 = d/"f1.csv"; f2 = d/"f2.xlsx"
    write_csv(f1, F1_CLIENTS); write_xlsx(f2, F2_REF)
    taille = os.path.getsize(f2); mtime = os.path.getmtime(f2)
    run_enrichissement(app, mod, f1, f2, d, "out.csv", "client_id", "cle_inexistante", MAPPING_BASE)
    assert any("ERREUR" in l for l in app._log_lines), "Aucun message ERREUR"
    assert os.path.getsize(f2) == taille,  "Taille F2 modifiée — INTERDIT"
    assert os.path.getmtime(f2) == mtime,  "Date F2 modifiée — INTERDIT"

# ══════════════════════════════════════════════════════════════════════════════
# GROUPE D — F1=XLSX, F2=XLSX
# ══════════════════════════════════════════════════════════════════════════════

@cas("D01-XLSX-F1_xlsx_F2_xlsx_standard")
def d01(app, mod):
    """F1=XLSX, F2=XLSX — 100% Excel. Aucun des deux fichiers modifié."""
    d = cas_dir("D01-XLSX-F1_xlsx_F2_xlsx_standard")
    f1 = d/"f1.xlsx"; f2 = d/"f2.xlsx"
    write_xlsx(f1, F1_CLIENTS); write_xlsx(f2, F2_REF)
    mf1 = os.path.getmtime(f1); mf2 = os.path.getmtime(f2)
    run_enrichissement(app, mod, f1, f2, d, "out.csv", "client_id", "sf_id", MAPPING_BASE)
    assert os.path.getmtime(f1) == mf1, "F1 xlsx modifié — INTERDIT"
    assert os.path.getmtime(f2) == mf2, "F2 xlsx modifié — INTERDIT"
    assert_enrichissement_ok(read_csv(d/"out.csv"))

@cas("D02-XLSX-F1_xlsx_F2_xlsx_rejet")
def d02(app, mod):
    """F1=XLSX, F2=XLSX — avec fichier de rejet. Attendu : résultat=5, rejet=1."""
    d = cas_dir("D02-XLSX-F1_xlsx_F2_xlsx_rejet")
    f1 = d/"f1.xlsx"; f2 = d/"f2.xlsx"
    write_xlsx(f1, F1_CLIENTS); write_xlsx(f2, F2_REF)
    run_enrichissement(app, mod, f1, f2, d, "out.csv", "client_id", "sf_id", MAPPING_BASE, rejet=True)
    ok = read_csv(d/"out.csv"); rej = read_csv(d/"out_Rejet.csv")
    assert len(ok) == 5, f"Enrichi attendu 5, obtenu {len(ok)}"
    assert len(rej) == 1, f"Rejet attendu 1, obtenu {len(rej)}"

@cas("D03-XLSX-F1_xlsx_F2_xlsx_mapping_multiple")
def d03(app, mod):
    """F1=XLSX, F2=XLSX — mapping 4 colonnes simultanées."""
    d = cas_dir("D03-XLSX-F1_xlsx_F2_xlsx_mapping_multiple")
    f1 = d/"f1.xlsx"; f2 = d/"f2.xlsx"
    write_xlsx(f1, F1_CLIENTS); write_xlsx(f2, F2_REF)
    run_enrichissement(app, mod, f1, f2, d, "out.csv", "client_id", "sf_id",
                       [("ville","city"),("code_postal","cp"),("contrat","contrat_type"),("segment","segment_client")])
    r = read_csv(d/"out.csv")
    assert r[0]["ville"] == "Paris",        f"C001 ville attendu Paris, obtenu '{r[0]['ville']}'"
    assert r[0]["segment"] == "Grand compte", f"C001 segment attendu 'Grand compte', obtenu '{r[0]['segment']}'"
    assert r[1]["contrat"] == "PME",        f"C002 contrat attendu PME, obtenu '{r[1]['contrat']}'"

@cas("D04-XLSX-F1_xlsx_F2_xlsx_structure_resultat_identique_F1")
def d04(app, mod):
    """F1=XLSX, F2=XLSX — header résultat identique à F1, aucune colonne de F2 ajoutée."""
    d = cas_dir("D04-XLSX-F1_xlsx_F2_xlsx_structure_resultat_identique_F1")
    f2_extra = [dict(r, **{"colonne_extra":"val"}) for r in F2_REF]
    f1 = d/"f1.xlsx"; f2 = d/"f2.xlsx"
    write_xlsx(f1, F1_CLIENTS); write_xlsx(f2, f2_extra)
    run_enrichissement(app, mod, f1, f2, d, "out.csv", "client_id", "sf_id", MAPPING_BASE)
    expected = list(F1_CLIENTS[0].keys())
    with open(d/"out.csv", "r", encoding="cp1252", errors="replace") as f:
        actual = next(csv.reader(f, delimiter=SEP))
    assert actual == expected, f"Header ≠ F1\n  F1  : {expected}\n  OUT : {actual}"

@cas("D05-XLSX-F1_xlsx_F2_xlsx_insensible_casse")
def d05(app, mod):
    """F1=XLSX, F2=XLSX — clé minuscule dans F2, mode insensible."""
    d = cas_dir("D05-XLSX-F1_xlsx_F2_xlsx_insensible_casse")
    f2_casse = [dict(r, **{"sf_id": r["sf_id"].lower()}) for r in F2_REF]
    f1 = d/"f1.xlsx"; f2 = d/"f2.xlsx"
    write_xlsx(f1, F1_CLIENTS); write_xlsx(f2, f2_casse)
    run_enrichissement(app, mod, f1, f2, d, "out.csv", "client_id", "sf_id", MAPPING_BASE, case_sensitive=False)
    assert read_csv(d/"out.csv")[0]["ville"] == "Paris"

@cas("D06-XLSX-F1_xlsx_F2_xlsx_aucun_fichier_modifie")
def d06(app, mod):
    """F1=XLSX, F2=XLSX — vérification exhaustive taille ET date des deux fichiers."""
    d = cas_dir("D06-XLSX-F1_xlsx_F2_xlsx_aucun_fichier_modifie")
    f1 = d/"f1.xlsx"; f2 = d/"f2.xlsx"
    write_xlsx(f1, F1_CLIENTS); write_xlsx(f2, F2_REF)
    tf1 = os.path.getsize(f1); mf1 = os.path.getmtime(f1)
    tf2 = os.path.getsize(f2); mf2 = os.path.getmtime(f2)
    run_enrichissement(app, mod, f1, f2, d, "out.csv", "client_id", "sf_id", MAPPING_BASE)
    assert os.path.getsize(f1)  == tf1, f"Taille F1 modifiée : {tf1}→{os.path.getsize(f1)}"
    assert os.path.getmtime(f1) == mf1, "Date F1 modifiée — INTERDIT"
    assert os.path.getsize(f2)  == tf2, f"Taille F2 modifiée : {tf2}→{os.path.getsize(f2)}"
    assert os.path.getmtime(f2) == mf2, "Date F2 modifiée — INTERDIT"

# ══════════════════════════════════════════════════════════════════════════════
# GROUPE E — CAS LIMITES EXCEL
# ══════════════════════════════════════════════════════════════════════════════

@cas("E01-XLSX-cellules_numeriques_converties_en_str")
def e01(app, mod):
    """F1=XLSX — entiers Excel → str sans '.0'. Attendu : '75001' pas '75001.0'."""
    d = cas_dir("E01-XLSX-cellules_numeriques_converties_en_str")
    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["ref","cp","nom","valeur"])
    ws.append([1001, 75001, "Paris Centre", 1.5])
    ws.append([1002, 69002, "Lyon", 2.0])
    f1 = d/"f1.xlsx"; wb.save(f1)
    write_csv(d/"f2.csv", [{"id":"1001","region":"IDF"},{"id":"1002","region":"ARA"}])
    run_enrichissement(app, mod, f1, d/"f2.csv", d, "out.csv", "ref", "id", [("nom","region")])
    r = read_csv(d/"out.csv")
    assert len(r) == 2,           f"Attendu 2 lignes, obtenu {len(r)}"
    assert r[0]["cp"] == "75001", f"75001 attendu, obtenu '{r[0]['cp']}'"
    assert "." not in r[0]["cp"], f"Pas de décimale : '{r[0]['cp']}'"

@cas("E02-XLSX-cellules_vides_dans_excel")
def e02(app, mod):
    """F1=XLSX — cellules None converties en '' (pas 'None').
    Attendu : enrichissement correct, nom != 'None'."""
    d = cas_dir("E02-XLSX-cellules_vides_dans_excel")
    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["client_id","nom","ville","contrat"])
    ws.append(["C001", None, None, None])
    ws.append(["C002", "Martin", None, None])
    f1 = d/"f1.xlsx"; wb.save(f1)
    write_csv(d/"f2.csv", F2_REF)
    run_enrichissement(app, mod, f1, d/"f2.csv", d, "out.csv", "client_id", "sf_id",
                       [("ville","city"),("contrat","contrat_type")])
    r = read_csv(d/"out.csv")
    assert r[0]["ville"]   == "Paris", f"C001 attendu Paris, obtenu '{r[0]['ville']}'"
    assert r[0]["nom"]     != "None",  f"Cellule vide ne doit pas être 'None', obtenu '{r[0]['nom']}'"
    assert r[1]["contrat"] == "PME",   f"C002 contrat attendu PME, obtenu '{r[1]['contrat']}'"

@cas("E03-XLSX-openpyxl_absent_erreur_explicite")
def e03(app, mod):
    """F1=XLSX — openpyxl absent : ERREUR explicite sans crash.
    Stratégie : patch le builtins.__import__ pour lever ImportError sur openpyxl."""
    import builtins
    d = cas_dir("E03-XLSX-openpyxl_absent_erreur_explicite")
    f1 = d/"f1.xlsx"; f2 = d/"f2.csv"
    write_xlsx(f1, F1_CLIENTS); write_csv(f2, F2_REF)

    real_import = builtins.__import__

    def fake_import(name, *args, **kwargs):
        if name == "openpyxl" or name.startswith("openpyxl."):
            raise ImportError("openpyxl non installé (simulé pour le test)")
        return real_import(name, *args, **kwargs)

    builtins.__import__ = fake_import
    try:
        run_enrichissement(app, mod, f1, f2, d, "out.csv", "client_id", "sf_id", MAPPING_BASE)
        errors = [l for l in app._log_lines if "ERREUR" in l]
        assert any("openpyxl" in l.lower() for l in errors), \
            f"ERREUR doit mentionner openpyxl. Log: {errors}"
    finally:
        builtins.__import__ = real_import

@cas("E04-XLSX-trim_espaces_cle_dans_xlsx")
def e04(app, mod):
    """F1=XLSX — clés avec espaces parasites. Attendu : strip() appliqué, match correct."""
    d = cas_dir("E04-XLSX-trim_espaces_cle_dans_xlsx")
    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["client_id","nom","ville","contrat"])
    ws.append(["  C001  ", "Dupont", "", ""])
    ws.append([" C002",    "Martin", "", ""])
    ws.append(["C003 ",    "Bernard","", ""])
    f1 = d/"f1.xlsx"; wb.save(f1)
    write_csv(d/"f2.csv", F2_REF)
    run_enrichissement(app, mod, f1, d/"f2.csv", d, "out.csv", "client_id", "sf_id",
                       [("ville","city"),("contrat","contrat_type")])
    r = read_csv(d/"out.csv")
    assert r[0]["ville"] == "Paris",     f"C001 attendu Paris, obtenu '{r[0]['ville']}'"
    assert r[1]["ville"] == "Lyon",      f"C002 attendu Lyon, obtenu '{r[1]['ville']}'"
    assert r[2]["ville"] == "Marseille", f"C003 attendu Marseille, obtenu '{r[2]['ville']}'"

# ══════════════════════════════════════════════════════════════════════════════
# GROUPE F — SORTIE EN XLSX
# ══════════════════════════════════════════════════════════════════════════════

@cas("F01-OUT_xlsx-F1_csv_F2_csv_sortie_xlsx")
def f01(app, mod):
    """F1=CSV, F2=CSV, sortie=XLSX — feuille 'Résultat' avec 6 lignes."""
    d = cas_dir("F01-OUT_xlsx-F1_csv_F2_csv_sortie_xlsx")
    f1 = d/"f1.csv"; f2 = d/"f2.csv"
    write_csv(f1, F1_CLIENTS); write_csv(f2, F2_REF)
    run_enrichissement(app, mod, f1, f2, d, "out.xlsx", "client_id", "sf_id", MAPPING_BASE, out_format="xlsx")
    assert (d/"out.xlsx").exists(), "Fichier out.xlsx non créé"
    assert_enrichissement_ok(read_xlsx(d/"out.xlsx", sheet="Résultat"))

@cas("F02-OUT_xlsx-rejet_dans_meme_classeur")
def f02(app, mod):
    """Sortie=XLSX avec rejet — feuille 'Rejet' dans le MÊME classeur.
    Attendu : Résultat=5, Rejet=1 (ZZZZ). Pas de fichier rejet séparé."""
    d = cas_dir("F02-OUT_xlsx-rejet_dans_meme_classeur")
    f1 = d/"f1.csv"; f2 = d/"f2.csv"
    write_csv(f1, F1_CLIENTS); write_csv(f2, F2_REF)
    run_enrichissement(app, mod, f1, f2, d, "out.xlsx", "client_id", "sf_id",
                       MAPPING_BASE, rejet=True, out_format="xlsx")
    assert (d/"out.xlsx").exists(),             "Fichier out.xlsx non créé"
    assert not (d/"out_Rejet.xlsx").exists(),   "Fichier rejet séparé ne doit PAS être créé"
    assert xlsx_sheet_exists(d/"out.xlsx", "Résultat"), "Feuille 'Résultat' absente"
    assert xlsx_sheet_exists(d/"out.xlsx", "Rejet"),    "Feuille 'Rejet' absente"
    r_ok  = read_xlsx(d/"out.xlsx", sheet="Résultat")
    r_rej = read_xlsx(d/"out.xlsx", sheet="Rejet")
    assert len(r_ok) == 5,  f"Résultat attendu 5, obtenu {len(r_ok)}"
    assert len(r_rej) == 1, f"Rejet attendu 1, obtenu {len(r_rej)}"
    assert r_rej[0]["client_id"] == "ZZZZ", f"Rejet attendu ZZZZ, obtenu '{r_rej[0]['client_id']}'"

@cas("F03-OUT_xlsx-F1_xlsx_F2_xlsx_sortie_xlsx")
def f03(app, mod):
    """F1=XLSX, F2=XLSX, sortie=XLSX — 100% Excel. Sources non modifiées."""
    d = cas_dir("F03-OUT_xlsx-F1_xlsx_F2_xlsx_sortie_xlsx")
    f1 = d/"f1.xlsx"; f2 = d/"f2.xlsx"
    write_xlsx(f1, F1_CLIENTS); write_xlsx(f2, F2_REF)
    mf1 = os.path.getmtime(f1); mf2 = os.path.getmtime(f2)
    run_enrichissement(app, mod, f1, f2, d, "out.xlsx", "client_id", "sf_id", MAPPING_BASE, out_format="xlsx")
    assert os.path.getmtime(f1) == mf1, "F1 modifié — INTERDIT"
    assert os.path.getmtime(f2) == mf2, "F2 modifié — INTERDIT"
    assert_enrichissement_ok(read_xlsx(d/"out.xlsx", sheet="Résultat"))

@cas("F04-OUT_xlsx-F1_xlsx_F2_csv_sortie_xlsx")
def f04(app, mod):
    """F1=XLSX, F2=CSV, sortie=XLSX — combinaison mixte."""
    d = cas_dir("F04-OUT_xlsx-F1_xlsx_F2_csv_sortie_xlsx")
    f1 = d/"f1.xlsx"; f2 = d/"f2.csv"
    write_xlsx(f1, F1_CLIENTS); write_csv(f2, F2_REF)
    run_enrichissement(app, mod, f1, f2, d, "out.xlsx", "client_id", "sf_id", MAPPING_BASE, out_format="xlsx")
    assert_enrichissement_ok(read_xlsx(d/"out.xlsx", sheet="Résultat"))

@cas("F05-OUT_xlsx-F1_csv_F2_xlsx_sortie_xlsx")
def f05(app, mod):
    """F1=CSV, F2=XLSX, sortie=XLSX — combinaison mixte."""
    d = cas_dir("F05-OUT_xlsx-F1_csv_F2_xlsx_sortie_xlsx")
    f1 = d/"f1.csv"; f2 = d/"f2.xlsx"
    write_csv(f1, F1_CLIENTS); write_xlsx(f2, F2_REF)
    run_enrichissement(app, mod, f1, f2, d, "out.xlsx", "client_id", "sf_id", MAPPING_BASE, out_format="xlsx")
    assert_enrichissement_ok(read_xlsx(d/"out.xlsx", sheet="Résultat"))

@cas("F06-OUT_xlsx-structure_header_identique_F1")
def f06(app, mod):
    """Sortie=XLSX — header feuille 'Résultat' identique à F1, colonnes F2 absentes."""
    d = cas_dir("F06-OUT_xlsx-structure_header_identique_F1")
    f2_extra = [dict(r, **{"colonne_extra":"val"}) for r in F2_REF]
    f1 = d/"f1.csv"; f2 = d/"f2.csv"
    write_csv(f1, F1_CLIENTS); write_csv(f2, f2_extra)
    run_enrichissement(app, mod, f1, f2, d, "out.xlsx", "client_id", "sf_id", MAPPING_BASE, out_format="xlsx")
    import openpyxl
    wb = openpyxl.load_workbook(d/"out.xlsx", read_only=True)
    ws = wb["Résultat"]
    header_out = [c.value for c in next(ws.iter_rows())]; wb.close()
    expected = list(F1_CLIENTS[0].keys())
    assert header_out == expected, f"Header xlsx ≠ F1\n  F1  : {expected}\n  OUT : {header_out}"

@cas("F07-OUT_xlsx-archivage_fichier_existant")
def f07(app, mod):
    """Sortie=XLSX — fichier existant archivé avec timestamp avant réécriture."""
    import time as _time
    d = cas_dir("F07-OUT_xlsx-archivage_fichier_existant")
    f1 = d/"f1.csv"; f2 = d/"f2.csv"
    write_csv(f1, F1_CLIENTS); write_csv(f2, F2_REF)
    run_enrichissement(app, mod, f1, f2, d, "out.xlsx", "client_id", "sf_id", MAPPING_BASE, out_format="xlsx")
    assert (d/"out.xlsx").exists(), "Premier fichier non créé"
    _time.sleep(1.1)
    run_enrichissement(app, mod, f1, f2, d, "out.xlsx", "client_id", "sf_id", MAPPING_BASE, out_format="xlsx")
    assert len(list(d.glob("out_2*.xlsx"))) >= 1, "Aucun fichier xlsx archivé trouvé"
    assert (d/"out.xlsx").exists(), "Nouveau fichier xlsx non créé"

# ══════════════════════════════════════════════════════════════════════════════
# RUNNER
# ══════════════════════════════════════════════════════════════════════════════

def merge_logs(global_log: Path):
    header = ["═"*62, f"  LOG GLOBAL — {time.strftime('%Y-%m-%d %H:%M:%S')}", "═"*62, ""]
    with open(global_log, "w", encoding="utf-8") as out:
        out.write("\n".join(header) + "\n")
        for lf in sorted(WORK_DIR.rglob("test.log")):
            try: out.write(lf.read_text(encoding="utf-8") + "\n")
            except Exception as e: out.write(f"  [warn] {lf} : {e}\n\n")

def run_all():
    src_file = demander_chemin_src()
    try:
        import openpyxl
    except ImportError:
        print(f"\n  {KO}  openpyxl non installé — requis pour les tests Excel.")
        print(f"       Exécutez : pip install openpyxl"); sys.exit(1)

    mod = _load_source(src_file)
    nb_ok = nb_ko = 0

    groupes = {
        "C": "Fonctionnalités CSV de base",
        "A": "CSV seul (référence formats)",
        "B": "F1=XLSX / F2=CSV",
        "D": "F1=XLSX / F2=XLSX",
        "E": "Cas limites Excel",
        "F": "Sortie en XLSX",
    }

    groupe_courant = None
    for nom, fn in TESTS:
        g = nom[0]
        if g != groupe_courant:
            groupe_courant = g
            label = groupes.get(g, g)
            print(f"\n  ── {label} ──")

        app = make_app(mod)
        erreur = ""; statut = "OK"; tb_str = ""
        try:
            fn(app, mod)
            print(f"  {OK}  {nom}"); nb_ok += 1
        except Exception as e:
            erreur = str(e); tb_str = _tb.format_exc(); statut = "KO"
            print(f"  {KO}  {nom}")
            print(f"       {INFO} {e}"); nb_ko += 1
        finally:
            if hasattr(app, "_run_config"):
                d = WORK_DIR / nom.replace(" ", "_").replace("—", "-")
                if d.exists():
                    try: write_log(d, nom, app, statut, erreur, tb_str)
                    except Exception as le: print(f"       [warn] log non écrit : {le}")

    global_log = Path(__file__).parent / f"tests_{time.strftime('%Y%m%d_%H%M%S')}.log"
    try:
        merge_logs(global_log)
        print(f"\n  {INFO}  Log global : {global_log}")
    except Exception as e:
        print(f"  [warn] Fusion des logs échouée : {e}")

    print("─"*62)
    total = nb_ok + nb_ko
    s = "\033[92m— TOUT OK\033[0m" if nb_ko == 0 else f"\033[91m— {nb_ko} ECHEC(S)\033[0m"
    print(f"  Résultat : {nb_ok}/{total} tests réussis  {s}")
    print(f"  Fichiers conservés dans : {WORK_DIR}")
    print("═"*62 + "\n")

    if sys.platform == "win32":
        import subprocess
        subprocess.Popen(["explorer", str(global_log.parent)])

    return nb_ko

if __name__ == "__main__":
    sys.exit(run_all())
